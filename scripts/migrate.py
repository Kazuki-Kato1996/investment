"""
売買表2026.xlsx → PostgreSQL マイグレーションスクリプト

使い方:
    pip install -r requirements.txt
    docker compose up -d
    python scripts/migrate.py

Docker経由:
    docker run --rm --network investment_default \
      -v $(pwd):/app -w /app -e DB_HOST=db \
      python:3.12-slim \
      sh -c "pip install -q openpyxl psycopg2-binary pandas && python scripts/migrate.py"
"""

import os
import re
from datetime import datetime, date

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', '売買表2026.xlsx')

DB_CONFIG = {
    'host':     os.environ.get('DB_HOST', 'localhost'),
    'port':     int(os.environ.get('DB_PORT', 5432)),
    'dbname':   os.environ.get('DB_NAME', 'investment_db'),
    'user':     os.environ.get('DB_USER', 'investor'),
    'password': os.environ.get('DB_PASSWORD', 'investment2026'),
}

# 月次シート一覧（年.月 形式）
MONTHLY_SHEETS = ['2026.1', '2026.2', '2026.3']


def to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def to_float(value) -> float | None:
    try:
        return float(value) if value is not None else None
    except (TypeError, ValueError):
        return None


def to_int(value) -> int | None:
    try:
        return int(value) if value is not None else None
    except (TypeError, ValueError):
        return None


def normalize_code(value) -> str | None:
    """銘柄コードを文字列に正規化（6674.0 → '6674'、'200A' → '200A'）"""
    if value is None:
        return None
    if isinstance(value, float):
        if value != int(value):
            return None  # 小数は無効
        return str(int(value))
    s = str(value).strip()
    # 数字のみ、または英数字混在コード（200A, 153A等）
    if re.match(r'^\d+[A-Z]?$', s):
        return s
    return None


def migrate_stocks(conn, wb):
    """銘柄マスタを購入価額シートから抽出"""
    print("銘柄マスタを移行中...")
    ws = wb['購入価額']
    stocks = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        # row[1]=コード, row[2]=銘柄名 の行を検出
        code = normalize_code(row[1]) if len(row) > 1 else None
        name = row[2] if len(row) > 2 else None
        if code and isinstance(name, str) and name not in ('配当利回り',):
            stocks[code] = name

    # 月次シートからも銘柄を補完
    for sheet_name in MONTHLY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws_m = wb[sheet_name]
        for row in ws_m.iter_rows(min_row=4, values_only=True):
            code = normalize_code(row[3]) if len(row) > 3 else None
            name = row[4] if len(row) > 4 else None
            if code and isinstance(name, str) and code not in stocks:
                stocks[code] = name

    records = [(c, n, 'JPY') for c, n in stocks.items()]

    with conn.cursor() as cur:
        execute_values(
            cur,
            """
            INSERT INTO stocks (code, name, currency)
            VALUES %s
            ON CONFLICT (code) DO UPDATE SET name = EXCLUDED.name
            """,
            records
        )
    conn.commit()
    print(f"  → {len(records)} 銘柄を登録")


def migrate_trades(conn, wb):
    """
    売買記録を月次シートから移行。
    各シートは証券会社セクションに分かれており、
    '売却日'を含む行がヘッダー、その後の行がトレードデータ。
    """
    print("売買記録を移行中...")
    trades = []

    for sheet_name in MONTHLY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        year_str, month_str = sheet_name.split('.')
        current_account = None

        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            if len(row) < 9:
                continue

            # アカウント名行を検出（例: '楽天証券', 'SBI証券', '岡地証券'）
            if row[4] and isinstance(row[4], str) and row[1] is None and row[3] is None:
                name = str(row[4]).strip()
                # 証券会社名っぽい行
                if any(kw in name for kw in ['証券', '銀行', 'SBI', '楽天', '岡地', '住信']):
                    current_account = name.replace('　', '').strip()

            # ヘッダー行をスキップ
            if row[1] == '売却日':
                continue

            # トレードデータ行: row[3]=コード, row[4]=銘柄名, row[6]=株数, row[7]=株価
            code = normalize_code(row[3]) if len(row) > 3 else None
            name = row[4] if len(row) > 4 else None
            quantity = to_int(to_float(row[6])) if len(row) > 6 else None
            price = to_float(row[7]) if len(row) > 7 else None
            trade_date = to_date(row[2]) if len(row) > 2 else None
            trade_type_raw = str(row[5]).strip() if len(row) > 5 and row[5] else ''

            if not code or not isinstance(name, str) or not quantity or not price:
                continue
            if quantity <= 0 or price <= 0:
                continue
            if trade_date is None:
                continue  # 日付なしはスキップ

            # 取引種別の正規化
            if '売' in trade_type_raw:
                trade_type = 'SELL'
            else:
                trade_type = 'BUY'

            trades.append((code, trade_date, trade_type, price, quantity, current_account))

    if trades:
        with conn.cursor() as cur:
            execute_values(
                cur,
                """
                INSERT INTO trades (code, trade_date, trade_type, price, quantity, account)
                VALUES %s
                ON CONFLICT DO NOTHING
                """,
                trades
            )
        conn.commit()
    print(f"  → {len(trades)} 件の売買記録を登録")


def migrate_monthly_snapshots(conn, wb):
    """月次スナップショット（月末時価・評価損益）を月次シートから移行"""
    print("月次スナップショットを移行中...")
    total = 0

    for sheet_name in MONTHLY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        year_val, month_val = map(int, sheet_name.split('.'))
        snapshot_month = date(year_val, month_val, 1)
        snapshots = []
        current_account = None

        for row in ws.iter_rows(values_only=True):
            if len(row) < 11:
                continue

            # アカウント名行
            if row[4] and isinstance(row[4], str) and row[1] is None and row[3] is None:
                name = str(row[4]).strip()
                if any(kw in name for kw in ['証券', '銀行', 'SBI', '楽天', '岡地', '住信']):
                    current_account = name.replace('　', '').strip()

            if row[1] == '売却日':
                continue

            code = normalize_code(row[3]) if len(row) > 3 else None
            if not code:
                continue

            market_value = to_float(row[10]) if len(row) > 10 else None   # 月末時価
            unrealized_gain = to_float(row[11]) if len(row) > 11 else None  # 評価損益
            purchase_price = to_float(row[8]) if len(row) > 8 else None     # 購入価格

            if market_value is None and unrealized_gain is None:
                continue

            # 現在株価を推算（月末時価 / 株数）
            quantity = to_int(to_float(row[6])) if len(row) > 6 else None
            current_price = None
            if market_value and quantity and quantity > 0:
                current_price = market_value / quantity

            snapshots.append((
                code, snapshot_month, current_price,
                market_value, unrealized_gain, None, current_account
            ))

        if snapshots:
            # (code, snapshot_month, account) の重複を除去（後勝ち）
            deduped = {}
            for s in snapshots:
                key = (s[0], s[1], s[6])
                deduped[key] = s
            snapshots = list(deduped.values())

            with conn.cursor() as cur:
                execute_values(
                    cur,
                    """
                    INSERT INTO monthly_snapshots
                        (code, snapshot_month, current_price, market_value, unrealized_gain, dividend, account)
                    VALUES %s
                    ON CONFLICT (code, snapshot_month, account) DO UPDATE SET
                        current_price   = EXCLUDED.current_price,
                        market_value    = EXCLUDED.market_value,
                        unrealized_gain = EXCLUDED.unrealized_gain
                    """,
                    snapshots
                )
            conn.commit()
        print(f"  → {sheet_name}: {len(snapshots)} 件")
        total += len(snapshots)

    print(f"  → 合計 {total} 件")


def migrate_annual_summaries(conn, wb):
    """
    年間集計シートを移行。
    構造:
      row2: ヘッダー（'2026年', '取得原価', '評価益', '時価評価額', '岡地現金', '住信SBI', 'SBI米株'）
      row3+: データ（月名, 取得原価, 評価益, 時価評価額, 口座別合計...）
    """
    print("年間集計を移行中...")
    ws = wb['年間集計']
    summaries = []
    current_year = None

    MONTH_MAP = {
        '1月': 1, '2月': 2, '3月': 3, '4月': 4,
        '5月': 5, '6月': 6, '7月': 7, '8月': 8,
        '9月': 9, '10月': 10, '11月': 11, '12月': 12,
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 年度行の検出（例: '2026年'）
        if isinstance(row[1], str) and '年' in str(row[1]):
            year_str = re.search(r'(\d{4})', str(row[1]))
            if year_str:
                current_year = int(year_str.group(1))
            continue

        if not current_year:
            continue

        month_str = str(row[1]).strip() if row[1] else ''
        month = MONTH_MAP.get(month_str)
        if not month:
            continue

        # 口座別残高（index 5=岡地現金, 6=住信SBI, 7=SBI米株）
        total_assets = to_float(row[4]) if len(row) > 4 else None  # 時価評価額

        for col_idx, account in [(5, '岡地現金'), (6, '住信SBI'), (7, 'SBI米株')]:
            val = to_float(row[col_idx]) if len(row) > col_idx else None
            if val is not None:
                summaries.append((account, current_year, month, None, None, None, val))

        # 合計も保存
        if total_assets is not None:
            summaries.append(('全口座合計', current_year, month, None,
                              to_float(row[3]) if len(row) > 3 else None,
                              None, total_assets))

    if summaries:
        # (account, year, month) の重複を除去
        deduped = {}
        for s in summaries:
            key = (s[0], s[1], s[2])
            deduped[key] = s
        summaries = list(deduped.values())

        with conn.cursor() as cur:
            execute_values(
                cur,
                """
                INSERT INTO annual_summaries
                    (account, year, month, cash, evaluation_gain, realized_gain, total_assets)
                VALUES %s
                ON CONFLICT (account, year, month) DO UPDATE SET
                    total_assets    = EXCLUDED.total_assets,
                    evaluation_gain = EXCLUDED.evaluation_gain
                """,
                summaries
            )
        conn.commit()
    print(f"  → {len(summaries)} 件の年間集計を登録")


def main():
    wb = load_workbook(XLSX_PATH)
    print(f"シート一覧: {wb.sheetnames}\n")

    conn = psycopg2.connect(**DB_CONFIG)
    try:
        migrate_stocks(conn, wb)
        migrate_trades(conn, wb)
        migrate_monthly_snapshots(conn, wb)
        migrate_annual_summaries(conn, wb)
        print("\n移行完了!")
    finally:
        conn.close()


def load_workbook(path: str):
    print(f"Excelファイルを読み込み中: {path}")
    return openpyxl.load_workbook(path, data_only=True)


if __name__ == '__main__':
    main()
